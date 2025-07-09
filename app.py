import streamlit as st
import pandas as pd
from openpyxl import load_workbook, Workbook
from io import BytesIO
import csv
import re
from typing import List, Tuple, Optional, Union, Dict, Any

# === Configurações iniciais ===
st.set_page_config(page_title="Agrupador de Entregas", page_icon="📦", layout="wide")

# === Tipos de dados ===
EnderecoLinha = List[Union[str, Optional[str]]]
TabelaProcessada = List[EnderecoLinha]
EntregaAgrupada = List[str]
DicionarioRuas = List[Tuple[str, str, str, str]]

# === Constantes ===
COLUNAS_ESPERADAS = 11
COLUNAS_OBRIGATORIAS = {3: "Número do Pacote", 10: "Bairro"}
NOME_ARQUIVO_DICIONARIO = "dicionario_ruas.csv"

# === Funções de processamento ===
def carregar_dicionario_ruas() -> DicionarioRuas:
    """Carrega o dicionário de ruas do arquivo CSV."""
    try:
        with open(NOME_ARQUIVO_DICIONARIO, encoding="utf-8") as csvfile:
            return [tuple(linha) for linha in csv.reader(csvfile)]
    except FileNotFoundError:
        st.error(f"⚠️ Arquivo {NOME_ARQUIVO_DICIONARIO} não encontrado. A correção de nomes de ruas não será aplicada.")
        return []
    except Exception as e:
        st.error(f"⚠️ Erro ao ler o arquivo {NOME_ARQUIVO_DICIONARIO}: {str(e)}")
        return []

def validar_linha(linha: List[Any], n_linha: int) -> Tuple[bool, Optional[str]]:
    """Valida uma linha da planilha."""
    if len(linha) < COLUNAS_ESPERADAS:
        return False, f"Linha {n_linha}: Número de colunas insuficiente (esperado {COLUNAS_ESPERADAS}, encontrado {len(linha)})"
    
    for col_idx, col_nome in COLUNAS_OBRIGATORIAS.items():
        if col_idx >= len(linha) or linha[col_idx].value is None:
            return False, f"Linha {n_linha}: Valor obrigatório vazio na coluna '{col_nome}'"
    
    return True, None

def gerar_df(uploaded_file: BytesIO) -> Tuple[TabelaProcessada, List[str]]:
    """Processa o arquivo Excel e gera a tabela de dados."""
    tabela = []
    erros = []
    dicionario_ruas = carregar_dicionario_ruas()
    
    try:
        wb = load_workbook(filename=uploaded_file, data_only=True)
        ws = wb.active
        
        if ws.max_row < 2:
            raise ValueError("O arquivo não contém dados suficientes (menos de 2 linhas)")
        
        for linha in ws.iter_rows(min_row=2):
            n_linha = linha[0].row
            linha_vals = [cell.value for cell in linha]
            
            # Verifica se a linha está oculta
            if ws.row_dimensions[n_linha].hidden:
                continue
                
            valido, msg_erro = validar_linha(linha, n_linha)
            if not valido:
                erros.append(msg_erro)
                continue
                
            try:
                # Extrai os valores necessários
                numero_pacote = str(linha[3].value).replace(".0", "") if linha[3].value is not None else ""
                endereco = str(linha[8].value) if linha[8].value is not None else ""
                bairro = str(linha[10].value) if linha[10].value is not None else ""
                
                tabela.append([numero_pacote, endereco, "", bairro])
            except Exception as e:
                erros.append(f"Linha {n_linha}: Erro ao processar - {str(e)}")
                
    except Exception as e:
        raise RuntimeError(f"Erro ao processar o arquivo: {str(e)}") from e
    
    return tabela, erros

def extrair_numero_endereco(endereco: str) -> Tuple[str, str]:
    """Extrai número do endereço e retorna rua e número separados."""
    if not endereco or not isinstance(endereco, str):
        return endereco, ""
        
    partes = endereco.split(',', 1)
    if len(partes) < 2:
        return endereco, ""
        
    rua = partes[0].strip()
    resto = partes[1].strip()
    match = re.match(r"^(sn|\d+)", resto, re.IGNORECASE)
    
    return rua, match.group(1).upper() if match else ""

def corrigir_sintaxe_ruas(tabela: TabelaProcessada) -> TabelaProcessada:
    """Corrige a sintaxe dos endereços, separando rua e número."""
    for linha in tabela:
        if len(linha) >= 2 and isinstance(linha[1], str):
            rua, numero = extrair_numero_endereco(linha[1])
            linha[1] = rua
            linha[2] = numero
    return tabela

def aplicar_correcoes_ruas(tabela: TabelaProcessada, dicionario: DicionarioRuas) -> TabelaProcessada:
    """Aplica correções de nomes de ruas baseadas no dicionário."""
    if not dicionario:
        return tabela
        
    correcoes_map = {(rua_errada, min_num, max_num): rua_correta 
                    for rua_errada, min_num, max_num, rua_correta in dicionario}
    
    for linha in tabela:
        if len(linha) < 3:
            continue
            
        rua = linha[1]
        numero = linha[2]
        
        for (rua_errada, min_num, max_num), rua_correta in correcoes_map.items():
            if rua == rua_errada and min_num <= numero <= max_num:
                linha[1] = rua_correta
                break
                
    return tabela

def agrupar_entregas(tabela: TabelaProcessada) -> List[EntregaAgrupada]:
    """Agrupa entregas por endereço."""
    if not tabela:
        return []
        
    # Ordena por rua, número e bairro
    tabela_ordenada = sorted(tabela, key=lambda x: (
        x[1] if len(x) > 1 and x[1] is not None else "", 
        x[2] if len(x) > 2 and x[2] is not None else "", 
        x[3] if len(x) > 3 and x[3] is not None else ""
    ))
    
    entregas_agrupadas = []
    grupo_atual = []
    endereco_atual = None
    
    for linha in tabela_ordenada:
        if len(linha) < 4:
            continue
            
        numero, rua, num_rua, bairro = linha
        endereco = (rua, num_rua, bairro)
        
        if endereco == endereco_atual:
            grupo_atual.append(numero)
        else:
            if grupo_atual and endereco_atual:
                entregas_agrupadas.append(formatar_entrega(grupo_atual, *endereco_atual))
            grupo_atual = [numero]
            endereco_atual = endereco
    
    if grupo_atual and endereco_atual:
        entregas_agrupadas.append(formatar_entrega(grupo_atual, *endereco_atual))
        
    return entregas_agrupadas

def formatar_entrega(pacotes: List[str], rua: str, numero: str, bairro: str) -> EntregaAgrupada:
    """Formata a saída de uma entrega agrupada."""
    if not pacotes:
        return ["", rua, numero, bairro]
        
    pacotes_str = pacotes[0] if len(pacotes) == 1 else ", ".join(pacotes[:-1]) + " e " + pacotes[-1]
    return [pacotes_str, rua, numero, bairro]

def gerar_planilha(tabela: List[EntregaAgrupada]) -> BytesIO:
    """Gera o arquivo Excel final."""
    buffer = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Entregas"
    ws.append(["Pacotes", "Rua", "Número", "Bairro"])
    
    for linha in tabela:
        ws.append(linha)
        
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def mostrar_erros(erros: List[str]) -> None:
    """Exibe erros de processamento para o usuário."""
    if not erros:
        return
        
    with st.expander("⚠️ Problemas encontrados (clique para ver detalhes)", expanded=False):
        st.warning(f"Foram encontrados {len(erros)} problemas durante o processamento:")
        for erro in erros:
            st.write(f"- {erro}")
        st.info("""
        **Soluções possíveis:**
        - Verifique se todas as colunas obrigatórias estão preenchidas
        - Confira se o formato do arquivo está correto
        - Linhas com problemas foram ignoradas no processamento
        """)

def mostrar_guia_uso() -> None:
    """Exibe um guia de uso da aplicação."""
    with st.expander("📌 Guia de Uso (clique para ver)", expanded=False):
        st.markdown("""
        **Como usar o Agrupador de Entregas:**
        
        1. **Prepare seu arquivo**: 
           - Deve ser um arquivo Excel (.xlsx)
           - Deve conter pelo menos as colunas obrigatórias:
             - Coluna 4: Número do pacote
             - Coluna 9: Endereço completo (rua, número)
             - Coluna 11: Bairro
        
        2. **Faça o upload** do arquivo usando o botão acima
        
        3. **Verifique os resultados**:
           - O sistema mostrará quantas paradas foram agrupadas
           - Se houver problemas, serão exibidos na seção de avisos
        
        4. **Baixe o resultado**:
           - Use o botão de download para obter a planilha agrupada
        
        **Dicas:**
        - Para melhores resultados, garanta que os endereços estejam no formato "Rua, Número"
        - O arquivo `dicionario_ruas.csv` pode ser usado para corrigir nomes de ruas automaticamente
        """)

# === Interface Streamlit ===
def main():
    st.title("📦 Agrupador de Entregas")
    st.markdown("Agrupe entregas por endereço para otimizar rotas de entrega")
    
    # mostrar_guia_uso()
    
    uploaded_file = st.file_uploader(
        "Selecione o arquivo Excel com as entregas", 
        type=["xlsx"],
        help="Arquivo deve conter colunas com número do pacote, endereço e bairro"
    )
    
    if uploaded_file:
        try:
            with st.spinner("Processando arquivo..."):
                tabela, erros = gerar_df(uploaded_file)
                
                if not tabela:
                    st.error("Nenhum dado válido encontrado no arquivo.")
                    return
                
                dicionario_ruas = carregar_dicionario_ruas()
                tabela_corrigida = corrigir_sintaxe_ruas(tabela)
                tabela_corrigida = aplicar_correcoes_ruas(tabela_corrigida, dicionario_ruas)
                entregas_agrupadas = agrupar_entregas(tabela_corrigida)
                
                # Mostrar resultados
                st.success(f"✅ Processamento concluído! A rota contém {len(entregas_agrupadas)} paradas.")
                
                # Mostrar pré-visualização
                with st.expander("🔍 Visualizar primeiras linhas", expanded=False):
                    preview_size = min(5, len(entregas_agrupadas))
                    st.table(pd.DataFrame(entregas_agrupadas[:preview_size], 
                                         columns=["Pacotes", "Rua", "Número", "Bairro"]))
                
                # Mostrar erros se houver
                # mostrar_erros(erros)
                
                # Botão de download
                arquivo_final = gerar_planilha(entregas_agrupadas)
                st.download_button(
                    label="📥 Baixar Planilha Agrupada",
                    data=arquivo_final,
                    file_name="entregas_agrupadas.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    help="Clique para baixar o arquivo com as entregas agrupadas por endereço"
                )
                
        except Exception as e:
            st.error(f"❌ Erro crítico ao processar o arquivo: {str(e)}")
            st.info("""
            **O que fazer agora?**
            - Verifique se o arquivo está no formato correto
            - Confira se todas as colunas obrigatórias estão presentes
            - Se o problema persistir, entre em contato com o suporte
            """)

if __name__ == "__main__":
    main()